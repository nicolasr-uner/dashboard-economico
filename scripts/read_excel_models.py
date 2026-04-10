"""
read_excel_models.py
Fase 3: Extrae métricas financieras de los modelos Excel de Exagon y Ruitoque
e inyecta los valores a la base de datos como ESTIMATION para el año base 2026.
"""
import sys, os
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl

from models.db import SessionLocal
from data.database import get_variables_by_name
from models.schema import TimeSeriesData, DataTypeEnum

BASE = r"c:\Users\Lenovo\OneDrive\Desktop\dashboard-economico"
EXAGON  = os.path.join(BASE, "Financial model - 24Mar2026 - Exagon - 13 Minifarms.xlsx")
RUITOQUE = os.path.join(BASE, "Financial model - Ruitoque Tax Partner.xlsx")

def update_db(metric_dict, model_label=""):
    session = SessionLocal()
    updated = 0
    for name, value in metric_dict.items():
        if value is None:
            continue
        try:
            val = float(value)
        except Exception:
            continue

        var_df = get_variables_by_name(name)
        if var_df.empty:
            print(f"  [SKIP] No en catalogo: {name}")
            continue

        var_id = int(var_df.iloc[0]['id'])
        record = TimeSeriesData(
            date=datetime(2026, 1, 1),
            variable_id=var_id,
            data_type=DataTypeEnum.ESTIMATION,
            value=val
        )
        session.merge(record)
        print(f"  [OK] {model_label} | {name} = {val}")
        updated += 1

    session.commit()
    session.close()
    return updated


def extract_from_excel(path, sheet_name):
    label = os.path.basename(path)
    print(f"\nLeyendo: {label} | Hoja: {sheet_name}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Error al abrir: {e}")
        return {}

    if sheet_name not in wb.sheetnames:
        print(f"  Hoja '{sheet_name}' no encontrada. Disponibles: {wb.sheetnames}")
        wb.close()
        return {}

    ws = wb[sheet_name]
    metrics = {
        "WACC - Costo Promedio de Capital": None,
        "Costo de la Deuda (Kd)": None,
        "Costo del Equity (Ke)": None,
        "Tarifa PPA (Precio Venta de Energ\u00eda)": None,
        "TIR Proyecto (IRR)": None,
        "CAPEX Solar Total (USD por proyecto)": None,
    }

    wacc_section = False
    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        if not any(c is not None for c in row[:8]):
            continue

        label_b = str(row[1]).strip() if row[1] is not None else ""
        label_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        val_d   = row[3] if len(row) > 3 else None
        val_h   = row[7] if len(row) > 7 else None

        # Detectar sección WACC
        if label_b.upper() == "WACC":
            wacc_section = True

        # WACC (valor fila ej. 'WACC', col D = 0.093557)
        if label_b == "WACC" and isinstance(val_d, (int, float)) and metrics["WACC - Costo Promedio de Capital"] is None:
            metrics["WACC - Costo Promedio de Capital"] = round(val_d * 100, 4)

        # Kd — row label = 'Kd', col D
        if label_b == "Kd" and isinstance(val_d, (int, float)):
            metrics["Costo de la Deuda (Kd)"] = round(val_d * 100, 4)

        # Ke — row label = 'Ke ' (con espacio), col D
        if label_b.startswith("Ke") and isinstance(val_d, (int, float)) and metrics["Costo del Equity (Ke)"] is None:
            metrics["Costo del Equity (Ke)"] = round(val_d * 100, 4)

        # IRR — en la tabla Results, col H (val_h) en fila que tiene 'IRR' y valor float
        if label_b == "IRR" and isinstance(val_h, (int, float)) and metrics["TIR Proyecto (IRR)"] is None:
            metrics["TIR Proyecto (IRR)"] = round(val_h * 100, 4)

        # PPA Price — '"PPA" Price' o 'PPA Price', col D number
        if ("PPA" in label_b and "Price" in label_b or
            ("PPA" in label_b and label_c == "COP")):
            if isinstance(val_d, (int, float)) and metrics["Tarifa PPA (Precio Venta de Energia)"] is None:
                metrics["Tarifa PPA (Precio Venta de Energia)"] = val_d

        # CAPEX total — 'Total Investment', col D USD value
        if "Total Investment" in label_b and isinstance(val_d, (int, float)):
            metrics["CAPEX Solar Total (USD por proyecto)"] = round(val_d, 2)

    wb.close()
    return metrics


if __name__ == "__main__":
    print("=" * 60)
    print("FASE 3: INGESTION EXCEL -> SQLITE")
    print("=" * 60)

    # --- EXAGON ---
    ex_metrics = extract_from_excel(EXAGON, "Assumptions & Results")
    print("\nExagon extraido:", {k: v for k, v in ex_metrics.items() if v is not None})
    ex_count = update_db(ex_metrics, "Exagon")
    print(f"  Exagon registros inyectados: {ex_count}")

    # --- RUITOQUE ---
    ru_metrics = extract_from_excel(RUITOQUE, "Assumptions & Results")
    print("\nRuitoque extraido:", {k: v for k, v in ru_metrics.items() if v is not None})
    # Ruitoque no pisa Exagon si ya hay valor; solo inserta lo faltante
    gaps = {k: v for k, v in ru_metrics.items() if v is not None and ex_metrics.get(k) is None}
    ru_count = update_db(gaps, "Ruitoque")
    print(f"  Ruitoque registros inyectados (adicionales): {ru_count}")

    print("\n--- FASE 3 COMPLETADA ---")
    print(f"Total inyectados: {ex_count + ru_count}")
